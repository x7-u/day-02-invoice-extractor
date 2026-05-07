"""Excel writer smoke tests -- workbook builds, sheets exist, cells are numeric."""
from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT.parent))

from excel_writer import write_workbook


@dataclass
class _R:
    invoice: dict
    source_filename: str
    cost_usd: float = 0.0
    input_tokens: int = 0
    output_tokens: int = 0
    model: str = ""
    n_pages: int = 1
    truncated: bool = False
    skipped: bool = False
    error: str | None = None


def _fake_invoice(currency: str = "GBP", confidence: str = "high"):
    return {
        "vendor": {"name": "Test Co", "vat_number": "GB123", "address": "1 Test St"},
        "invoice_number": "INV-1",
        "invoice_date": "2026-04-01",
        "due_date": "2026-05-01",
        "currency": currency,
        "line_items": [
            {"description": "Item 1", "quantity": 1, "unit_price": 50.0, "total": 50.0},
            {"description": "Item 2", "quantity": 2, "unit_price": 25.0, "total": 50.0},
        ],
        "subtotal": 100.0,
        "tax_amount": 20.0,
        "tax_rate": 0.20,
        "total_amount": 120.0,
        "confidence": confidence,
        "extraction_warnings": [],
    }


def test_workbook_has_three_sheets(tmp_path):
    out = write_workbook([_R(invoice=_fake_invoice(), source_filename="a.pdf")], tmp_path)
    wb = load_workbook(out)
    names = wb.sheetnames
    assert "Invoices" in names
    assert "Line items" in names
    assert "Summary" in names


def test_invoices_sheet_has_numeric_totals(tmp_path):
    out = write_workbook([_R(invoice=_fake_invoice(), source_filename="a.pdf", cost_usd=0.005)], tmp_path)
    wb = load_workbook(out)
    ws = wb["Invoices"]
    # Row 2 column 9 (Total) should be a number.
    assert isinstance(ws.cell(row=2, column=9).value, (int, float))
    assert ws.cell(row=2, column=9).value == 120.0


def test_line_items_sheet_one_row_per_line(tmp_path):
    out = write_workbook([_R(invoice=_fake_invoice(), source_filename="a.pdf")], tmp_path)
    wb = load_workbook(out)
    ws = wb["Line items"]
    # Header row + 2 line items
    assert ws.max_row == 3


def test_summary_aggregates_by_currency(tmp_path):
    invs = [
        _R(invoice=_fake_invoice("GBP"), source_filename="a.pdf", cost_usd=0.005),
        _R(invoice=_fake_invoice("EUR"), source_filename="b.pdf", cost_usd=0.005),
        _R(invoice=_fake_invoice("GBP"), source_filename="c.pdf", cost_usd=0.005),
    ]
    out = write_workbook(invs, tmp_path)
    wb = load_workbook(out)
    ws = wb["Summary"]
    # Find the currency rows starting at row 11 (per the layout in excel_writer.py).
    currencies_seen = set()
    for r in range(11, 16):
        v = ws.cell(row=r, column=1).value
        if v:
            currencies_seen.add(v)
    assert "GBP" in currencies_seen
    assert "EUR" in currencies_seen


def test_filename_includes_timestamp(tmp_path):
    out = write_workbook([_R(invoice=_fake_invoice(), source_filename="a.pdf")], tmp_path)
    assert out.name.startswith("invoices_")
    assert out.name.endswith(".xlsx")
    # Format: invoices_YYYYMMDD-HHMM.xlsx
    parts = out.stem.split("_")
    assert len(parts) == 2
    ts = parts[1]
    assert "-" in ts and len(ts) >= 13
