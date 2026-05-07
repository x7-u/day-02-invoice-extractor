"""Excel output -- three sheets: Invoices, Line items, Summary.

Invoices: one row per invoice, summary fields, hyperlink-style "see lines" cue.
Line items: one row per line, with foreign-key invoice_number to join.
Summary: count by vendor, total spend by currency, count by confidence band,
embedded bar chart of confidence distribution.

Reuses the RAG palette from shared/chart_styles for the confidence colouring.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.chart_styles import RAG_HEX

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

CONFIDENCE_FILL = {
    "high":   PatternFill("solid", fgColor=RAG_HEX["green"]),
    "medium": PatternFill("solid", fgColor=RAG_HEX["amber"]),
    "low":    PatternFill("solid", fgColor=RAG_HEX["red"]),
}

DUP_FILL = {
    "exact":      PatternFill("solid", fgColor=RAG_HEX["red"]),
    "suspicious": PatternFill("solid", fgColor=RAG_HEX["amber"]),
    "fuzzy":      PatternFill("solid", fgColor="7B2CBF"),  # plum
}


def _money_format(currency: str | None) -> str:
    """Excel format string for the currency. £/€/$ glyphs render correctly."""
    sym = {"GBP": "£", "EUR": "€", "USD": "$"}.get((currency or "").upper(), "")
    if sym:
        return f'"{sym}"#,##0.00;[Red]-"{sym}"#,##0.00'
    return "#,##0.00;[Red]-#,##0.00"


def _write_invoices_sheet(ws, results: list) -> None:
    ws.title = "Invoices"
    headers = [
        "Source file", "Vendor", "Invoice #", "Invoice date", "Due date",
        "Currency", "Subtotal", "Tax", "Total", "# Line items", "Confidence",
        "Duplicate", "Matches", "Warnings",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    for r_idx, r in enumerate(results, start=2):
        inv = r.invoice
        vendor_name = (inv.get("vendor") or {}).get("name") or ""
        warnings = "; ".join(inv.get("extraction_warnings") or [])
        currency = inv.get("currency") or ""
        money_fmt = _money_format(currency)

        ws.cell(row=r_idx, column=1, value=r.source_filename)
        ws.cell(row=r_idx, column=2, value=vendor_name)
        ws.cell(row=r_idx, column=3, value=inv.get("invoice_number") or "")
        ws.cell(row=r_idx, column=4, value=inv.get("invoice_date") or "")
        ws.cell(row=r_idx, column=5, value=inv.get("due_date") or "")
        ws.cell(row=r_idx, column=6, value=currency)
        for col, key in [(7, "subtotal"), (8, "tax_amount"), (9, "total_amount")]:
            v = inv.get(key)
            cell = ws.cell(row=r_idx, column=col, value=v)
            if v is not None:
                cell.number_format = money_fmt
        ws.cell(row=r_idx, column=10, value=len(inv.get("line_items") or []))
        conf_cell = ws.cell(row=r_idx, column=11, value=(inv.get("confidence") or "").upper())
        conf_cell.alignment = Alignment(horizontal="center")
        conf_cell.font = Font(bold=True, color="FFFFFF")
        conf_fill = CONFIDENCE_FILL.get((inv.get("confidence") or "").lower())
        if conf_fill:
            conf_cell.fill = conf_fill

        dup_status = getattr(r, "duplicate_status", None) or ""
        dup_match = getattr(r, "duplicate_match_filename", None) or ""
        dup_cell = ws.cell(row=r_idx, column=12, value=dup_status.upper() if dup_status else "")
        dup_cell.alignment = Alignment(horizontal="center")
        if dup_status in DUP_FILL:
            dup_cell.fill = DUP_FILL[dup_status]
            dup_cell.font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r_idx, column=13, value=dup_match)
        ws.cell(row=r_idx, column=14, value=warnings)

    widths = [22, 32, 18, 14, 14, 10, 14, 14, 14, 12, 12, 12, 24, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_line_items_sheet(ws, results: list) -> None:
    ws.title = "Line items"
    headers = [
        "Source file", "Invoice #", "Vendor", "Currency",
        "Description", "Quantity", "Unit price", "Line total",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    row = 2
    for r in results:
        inv = r.invoice
        vendor = (inv.get("vendor") or {}).get("name") or ""
        currency = inv.get("currency") or ""
        money_fmt = _money_format(currency)
        for li in inv.get("line_items") or []:
            ws.cell(row=row, column=1, value=r.source_filename)
            ws.cell(row=row, column=2, value=inv.get("invoice_number") or "")
            ws.cell(row=row, column=3, value=vendor)
            ws.cell(row=row, column=4, value=currency)
            ws.cell(row=row, column=5, value=li.get("description") or "")
            qty_cell = ws.cell(row=row, column=6, value=li.get("quantity"))
            if li.get("quantity") is not None:
                qty_cell.number_format = "0.##"
            up_cell = ws.cell(row=row, column=7, value=li.get("unit_price"))
            if li.get("unit_price") is not None:
                up_cell.number_format = money_fmt
            tot_cell = ws.cell(row=row, column=8, value=li.get("total"))
            if li.get("total") is not None:
                tot_cell.number_format = money_fmt
            row += 1

    widths = [22, 18, 32, 10, 50, 10, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary_sheet(ws, results: list) -> None:
    ws.title = "Summary"
    ws["A1"] = f"Invoice extraction summary -- {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    # Confidence distribution
    counts = {"high": 0, "medium": 0, "low": 0}
    for r in results:
        c = (r.invoice.get("confidence") or "").lower()
        if c in counts:
            counts[c] += 1

    ws["A3"] = "Confidence distribution"
    ws["A3"].font = Font(bold=True)
    headers = ["Band", "Count"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, (band, n) in enumerate(counts.items(), start=5):
        cell = ws.cell(row=i, column=1, value=band.upper())
        cell.fill = CONFIDENCE_FILL[band]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=n)

    # Total spend by currency
    by_currency: dict[str, float] = {}
    for r in results:
        inv = r.invoice
        cur = inv.get("currency") or "(unknown)"
        amt = inv.get("total_amount") or 0
        by_currency[cur] = by_currency.get(cur, 0) + amt

    ws["A9"] = "Total extracted spend by currency"
    ws["A9"].font = Font(bold=True)
    for col_idx, h in enumerate(["Currency", "Total"], start=1):
        cell = ws.cell(row=10, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, (cur, total) in enumerate(by_currency.items(), start=11):
        ws.cell(row=i, column=1, value=cur)
        cell = ws.cell(row=i, column=2, value=total)
        cell.number_format = _money_format(cur)

    # Aggregate cost
    total_cost = sum(r.cost_usd for r in results)
    ws.cell(row=20, column=1, value="Total Claude API cost (USD):").font = Font(bold=True)
    cell = ws.cell(row=20, column=2, value=total_cost)
    cell.number_format = '"$"#,##0.0000'

    # Embedded bar chart of confidence distribution
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Invoices by confidence band"
    chart.y_axis.title = "Band"
    chart.x_axis.title = "Count"
    data = Reference(ws, min_col=2, min_row=4, max_row=7)
    cats = Reference(ws, min_col=1, min_row=5, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D3")

    widths = [28, 16, 4, 4, 4, 4]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_workbook(results: list, out_dir: Path) -> Path:
    """Build the three-sheet workbook and write it to a versioned filename."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M")
    out = out_dir / f"invoices_{ts}.xlsx"

    wb = Workbook()
    _write_invoices_sheet(wb.active, results)
    _write_line_items_sheet(wb.create_sheet("Line items"), results)
    _write_summary_sheet(wb.create_sheet("Summary"), results)
    wb.save(out)
    return out
